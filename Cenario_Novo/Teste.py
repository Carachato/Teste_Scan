import os
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

pastas = ["CPY", "CP2", "CDD", "SRC"]
campos = ["CIC", "CPF", "CNPJ", "CGC", "FILIAL", "CNINTATR", "DOC"]
saida = "resultado_busca.xlsx"

def calcular_tamanho_pic(pic_str: str) -> int:
    pic_str = pic_str.strip()
    pic_str = re.sub(r"\s+VALUE\s+.*", "", pic_str)
    if "." in pic_str:
        pic_str = pic_str.split(".")[0].strip()
    match = re.match(r"PIC\s+[SX9]\((\d+)\)", pic_str)
    if match: return int(match.group(1))
    match = re.match(r"PIC\s+S?9\((\d+)\)V9\((\d+)\)", pic_str)
    if match: return int(match.group(1)) + int(match.group(2))
    match = re.match(r"PIC\s+S?9\((\d+)\).*COMP-3", pic_str)
    if match: return (int(match.group(1)) + 2) // 2
    return 0

estatisticas = {pasta: {"arquivos": 0, "registros": 0} for pasta in pastas}
total_registros = 0
tempo_inicial = datetime.now()

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

# Criar abas para cada pasta
for pasta in pastas:
    ws = wb.create_sheet(title=pasta)
    ws.append(["Tipo", "Módulo", "Campo Encontrado", "Tipo Campo", "Existe", "Origem"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    if os.path.exists(pasta):
        for root, dirs, files in os.walk(pasta):
            for file in files:
                estatisticas[pasta]["arquivos"] += 1
                caminho_arquivo = os.path.join(root, file)
                modulo = os.path.splitext(file)[0]
                registros_encontrados = set()
                try:
                    with open(caminho_arquivo, "r", encoding="utf-8", errors="ignore") as f:
                        for linha in f:
                            linha_upper = linha.upper()
                            if len(linha_upper) >= 7 and linha_upper[6] == "*": continue
                            if len(linha_upper) >= 72 and linha_upper[71] == "*": continue
                            if "SQLCA" in linha_upper or "SC5LDIS1" in linha_upper or "SC5LDIS2" in linha_upper: continue
                            linha_cortada = linha_upper[7:73].strip()
                            if pasta == "SRC":
                                partes = linha_cortada.split()
                                if "INCLUDE" in partes:
                                    idx = partes.index("INCLUDE")
                                    if idx+1 < len(partes): registros_encontrados.add(("INCLUDE", partes[idx+1], "Diretiva"))
                                elif "$COPY" in partes:
                                    idx = partes.index("$COPY")
                                    if idx+1 < len(partes): registros_encontrados.add(("$COPY", partes[idx+1], "Diretiva"))
                                elif "COPY" in partes:
                                    idx = partes.index("COPY")
                                    if idx+1 < len(partes): registros_encontrados.add(("COPY", partes[idx+1], "Diretiva"))
                            if any(campo in linha_cortada for campo in campos):
                                linha_sem_nivel = re.sub(r"^\s*\d+\s+", "", linha_cortada)
                                linha_formatada = re.sub(r"\s+", " ", linha_sem_nivel)
                                match = re.match(r"([A-Z0-9\-]+)\s+(PIC.*)", linha_formatada)
                                if match:
                                    nome_var = match.group(1)
                                    tipo_pic = match.group(2)
                                    tipo_pic = re.sub(r"\s+VALUE\s+.*", "", tipo_pic)
                                    if "." in tipo_pic: tipo_pic = tipo_pic.split(".")[0].strip()
                                    tamanho = calcular_tamanho_pic(tipo_pic)
                                    if tamanho > 0 and tamanho <= 18:
                                        registros_encontrados.add((nome_var, tipo_pic, "Variável"))
                    if registros_encontrados:
                        for nome_var, tipo_pic, origem in registros_encontrados:
                            ws.append([pasta, modulo, nome_var, tipo_pic, "X", origem])
                            estatisticas[pasta]["registros"] += 1; total_registros += 1
                    else:
                        ws.append([pasta, modulo, "-", "-", "N", "-"])
                        estatisticas[pasta]["registros"] += 1; total_registros += 1
                except Exception:
                    ws.append([pasta, modulo, "Erro ao abrir", "-", "N", "-"])
                    estatisticas[pasta]["registros"] += 1; total_registros += 1

# Aba review
ws_review = wb.create_sheet(title="review")
ws_review.append(["Módulo", "Campo Encontrado", "Tipo Campo", "Notificação", "Origem"])
for cell in ws_review[1]:
    cell.font = header_font; cell.fill = header_fill

ws_src = wb["SRC"]
for row in ws_src.iter_rows(min_row=2, values_only=True):
    modulo_src, campo_src, tipo_src, existe_src, origem_src = row[1], row[2], row[3], row[4], row[5]
    if origem_src == "Variável":
        if "PIC 9" in tipo_src or "PIC S9" in tipo_src:
            ws_review.append([modulo_src, campo_src, tipo_src, "ponto de atenção para o uso da variavél", origem_src])
    elif origem_src == "Diretiva":
        ref_copy = tipo_src.strip()
        for aba in ["CPY", "CDD", "CP2"]:
            ws_check = wb[aba]
            for row_check in ws_check.iter_rows(min_row=2, values_only=True):
                modulo_chk, campo_chk, tipo_chk, existe_chk, origem_chk = row_check[1], row_check[2], row_check[3], row_check[4], row_check[5]
                if modulo_chk.strip() == ref_copy:
                    if existe_chk == "X":
                        ws_review.append([modulo_src, campo_chk, tipo_chk, "ponto de atenção para o uso da variavél", origem_src])

# Aba Resumo
ws_resumo = wb.create_sheet(title="Resumo")
ws_resumo.append(["Pasta", "Arquivos Lidos", "Registros Gerados"])
for cell in ws_resumo[1]:
    cell.font = header_font; cell.fill = header_fill
for pasta, dados in estatisticas.items():
    ws_resumo.append([pasta, dados["arquivos"], dados["registros"]])
ws_resumo.append(["TOTAL", sum(d["arquivos"] for d in estatisticas.values()), total_registros])

wb.save(saida)

tempo_final = datetime.now()
tempo_decorrido = tempo_final - tempo_inicial

print(f"Arquivo Excel '{saida}' gerado com sucesso!\n")
print("📊 Estatísticas da execução:")
for pasta, dados in estatisticas.items():
    print(f"- {pasta}: {dados['arquivos']} arquivos lidos, {dados['registros']} registros gerados")
print(f"➡️ Total de registros no Excel: {total_registros}")
print("\n⏱️ Tempo de execução:")
print(f"- Início: {tempo_inicial}")
print(f"- Fim:    {tempo_final}")
print(f"- Decorrido: {tempo_decorrido}")
